"""Shared atomic write-back utilities for CSV and Excel sources."""

from __future__ import annotations

import csv
import os
import tempfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for excel flows
    load_workbook = None  # type: ignore[assignment]


def _cleanup_temp_artifacts(temp_fd: int | None, temp_path: str | None) -> None:
    """Best-effort cleanup for temporary file descriptor and path."""
    if temp_fd is not None:
        try:
            os.close(temp_fd)
        except OSError:
            pass
    if temp_path is not None and os.path.exists(temp_path):
        try:
            os.unlink(temp_path)
        except OSError:
            pass


def apply_csv_updates_atomic(
    file_path: str,
    row_updates: dict[int, dict[str, Any]],
) -> int:
    """Apply row updates to a CSV file with temp-file + atomic replace semantics.

    Args:
        file_path: Absolute/relative CSV file path.
        row_updates: Mapping of 1-based row number -> column/value updates.

    Returns:
        Number of updated rows.

    Raises:
        ValueError: If the CSV is empty or a row number is out of range.
    """
    if not row_updates:
        return 0

    with open(file_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    if not fieldnames:
        raise ValueError("CSV file has no header row.")

    if not rows:
        raise ValueError("CSV file has no data rows.")

    for row_number in row_updates:
        if row_number < 1 or row_number > len(rows):
            raise ValueError(
                f"Row {row_number} not found. CSV has {len(rows)} data rows.",
            )

    needed_columns = set(fieldnames)
    for updates in row_updates.values():
        needed_columns.update(updates.keys())

    ordered_fieldnames = fieldnames + [
        col for col in needed_columns if col not in fieldnames
    ]

    for row_number, updates in row_updates.items():
        row = rows[row_number - 1]
        for column, value in updates.items():
            row[column] = "" if value is None else str(value)

    dir_path = str(Path(file_path).resolve().parent)
    temp_fd: int | None = None
    temp_path: str | None = None
    try:
        temp_fd, temp_path = tempfile.mkstemp(suffix=".csv.tmp", dir=dir_path)
        with os.fdopen(temp_fd, "w", newline="", encoding="utf-8") as f:
            temp_fd = None
            writer = csv.DictWriter(f, fieldnames=ordered_fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temp_path, file_path)
        temp_path = None
    except Exception:
        _cleanup_temp_artifacts(temp_fd, temp_path)
        raise

    return len(row_updates)


def apply_delimited_updates_atomic(
    file_path: str,
    row_updates: dict[int, dict[str, Any]],
    delimiter: str = ",",
) -> int:
    """Apply row updates to a delimited file preserving the original delimiter.

    Same logic as apply_csv_updates_atomic but with configurable delimiter.

    Args:
        file_path: Absolute/relative delimited file path.
        row_updates: Mapping of 1-based row number -> column/value updates.
        delimiter: Column delimiter (default: comma).

    Returns:
        Number of updated rows.

    Raises:
        ValueError: If the file has no header, no data rows, or row number is out of range.
    """
    if not row_updates:
        return 0

    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError(f"File has no header row: {file_path}")
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    if not rows:
        raise ValueError(f"File has no data rows: {file_path}")

    # Collect any new columns needed
    needed_columns: set[str] = set()
    for updates in row_updates.values():
        needed_columns.update(updates.keys())
    ordered_fieldnames = fieldnames + [
        col for col in sorted(needed_columns) if col not in fieldnames
    ]

    updated_count = 0
    for row_number, updates in row_updates.items():
        if row_number < 1 or row_number > len(rows):
            raise ValueError(
                f"Row {row_number} out of range (1-{len(rows)})"
            )
        row = rows[row_number - 1]
        for column, value in updates.items():
            row[column] = "" if value is None else str(value)
        updated_count += 1

    # Atomic write via temp file + rename
    dir_path = str(Path(file_path).resolve().parent)
    temp_fd: int | None = None
    temp_path: str | None = None
    try:
        temp_fd, temp_path = tempfile.mkstemp(suffix=".tmp", dir=dir_path)
        with os.fdopen(temp_fd, "w", newline="", encoding="utf-8") as f:
            temp_fd = None
            writer = csv.DictWriter(
                f, fieldnames=ordered_fieldnames, delimiter=delimiter
            )
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temp_path, file_path)
        temp_path = None
    except Exception:
        _cleanup_temp_artifacts(temp_fd, temp_path)
        raise

    return updated_count


def write_companion_csv(
    source_path: str,
    row_number: int,
    reference_id: str,
    tracking_number: str,
    shipped_at: str,
    cost_cents: int | None = None,
) -> str:
    """Write a row to the companion results CSV for non-flat source formats.

    Creates {source_stem}_results.csv on first call; appends on subsequent calls.
    Used for JSON, XML, EDI, and other hierarchical formats where modifying
    the original file would be structurally destructive.

    Args:
        source_path: Path to the original source file (for deriving companion path).
        row_number: 1-based row number from imported_data.
        reference_id: Order reference ID (for human identification).
        tracking_number: UPS tracking number.
        shipped_at: ISO8601 timestamp.
        cost_cents: Shipping cost in cents (optional).

    Returns:
        Path to the companion CSV file.
    """
    source = Path(source_path)
    companion_path = source.parent / f"{source.stem}_results.csv"

    fieldnames = [
        "Original_Row_Number",
        "Reference_ID",
        "Tracking_Number",
        "Shipped_At",
        "Cost_Cents",
    ]
    row_data = {
        "Original_Row_Number": row_number,
        "Reference_ID": reference_id,
        "Tracking_Number": tracking_number,
        "Shipped_At": shipped_at,
        "Cost_Cents": cost_cents or "",
    }

    write_header = not companion_path.exists()
    with open(companion_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow(row_data)

    return str(companion_path)


def apply_excel_updates_atomic(
    file_path: str,
    row_updates: dict[int, dict[str, Any]],
    sheet_name: str | None = None,
) -> int:
    """Apply row updates to an Excel file with atomic replace semantics.

    Args:
        file_path: Absolute/relative Excel file path.
        row_updates: Mapping of 1-based data row number -> column/value updates.
        sheet_name: Optional target sheet; defaults to active sheet.

    Returns:
        Number of updated rows.

    Raises:
        ImportError: If openpyxl is unavailable.
        ValueError: If worksheet/header/row validation fails.
    """
    if not row_updates:
        return 0
    if load_workbook is None:
        raise ImportError("openpyxl is required for Excel write-back")

    wb = load_workbook(file_path)
    temp_fd: int | None = None
    temp_path: str | None = None
    try:
        if sheet_name and sheet_name != "(first sheet)":
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            raise ValueError("No active worksheet found")

        headers = [cell.value for cell in ws[1]]
        if not headers:
            raise ValueError("Excel sheet has no header row.")

        header_index: dict[str, int] = {}
        for idx, header in enumerate(headers, start=1):
            if isinstance(header, str) and header:
                header_index[header] = idx

        needed_columns: list[str] = []
        for updates in row_updates.values():
            for column in updates:
                if column not in header_index and column not in needed_columns:
                    needed_columns.append(column)

        next_col = len(headers) + 1
        for column in needed_columns:
            ws.cell(row=1, column=next_col, value=column)
            header_index[column] = next_col
            next_col += 1

        max_data_rows = max(0, ws.max_row - 1)
        for row_number in row_updates:
            if row_number < 1 or row_number > max_data_rows:
                raise ValueError(
                    f"Row {row_number} not found. Excel has {max_data_rows} data rows.",
                )

        for row_number, updates in row_updates.items():
            excel_row = row_number + 1
            for column, value in updates.items():
                ws.cell(row=excel_row, column=header_index[column], value=value)

        dir_path = str(Path(file_path).resolve().parent)
        temp_fd, temp_path = tempfile.mkstemp(
            suffix=f"{Path(file_path).suffix}.tmp",
            dir=dir_path,
        )
        os.close(temp_fd)
        temp_fd = None
        wb.save(temp_path)
        wb.close()
        os.replace(temp_path, file_path)
        temp_path = None
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        _cleanup_temp_artifacts(temp_fd, temp_path)
        raise

    return len(row_updates)
