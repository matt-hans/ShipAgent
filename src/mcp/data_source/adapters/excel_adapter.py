"""Excel adapter for Data Source MCP.

Provides Excel file import capabilities using openpyxl for .xlsx sheet discovery
and data reading, with python-calamine for legacy .xls files and DuckDB for SQL querying.

Per CONTEXT.md:
- One source at a time (importing replaces previous)
- Silent skip for empty rows
- Best-effort parsing with string fallback
"""

from pathlib import Path
from typing import TYPE_CHECKING, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import BaseSourceAdapter
from ..models import SOURCE_ROW_NUM_COLUMN, ImportResult, SchemaColumn
from ..utils import parse_date_with_warnings

try:
    from python_calamine import CalamineWorkbook

    _calamine_available = True
except ImportError:
    _calamine_available = False

if TYPE_CHECKING:
    from duckdb import DuckDBPyConnection


class ExcelAdapter(BaseSourceAdapter):
    """Adapter for importing Excel files via openpyxl and DuckDB.

    Uses openpyxl for:
    - Sheet discovery (list_sheets)
    - Reading cell data with proper type handling

    Uses DuckDB for:
    - SQL access to imported data
    - Schema storage and querying
    """

    @property
    def source_type(self) -> str:
        """Return the adapter's source type identifier.

        Returns:
            Source type string: 'excel'
        """
        return "excel"

    def _is_legacy_xls(self, file_path: str) -> bool:
        """Check if file is legacy .xls format.

        Args:
            file_path: Path to the Excel file.

        Returns:
            True if the file extension is .xls (case-insensitive).
        """
        return Path(file_path).suffix.lower() == ".xls"

    def list_sheets(self, file_path: str) -> list[str]:
        """List all sheet names in an Excel file.

        Uses openpyxl for .xlsx, python-calamine for legacy .xls.

        Args:
            file_path: Path to Excel file (.xlsx or .xls)

        Returns:
            List of sheet names in workbook order

        Raises:
            FileNotFoundError: If file does not exist
            ImportError: If .xls file and python-calamine not installed
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if self._is_legacy_xls(file_path):
            return self._list_sheets_calamine(file_path)

        # Use read_only mode for efficiency - only loads sheet metadata
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        return sheet_names

    def _list_sheets_calamine(self, file_path: str) -> list[str]:
        """List sheets using python-calamine for legacy .xls files.

        Args:
            file_path: Path to .xls file.

        Returns:
            List of sheet names.

        Raises:
            ImportError: If python-calamine is not installed.
        """
        if not _calamine_available:
            raise ImportError(
                "python-calamine required for .xls files: "
                "pip install python-calamine"
            )
        wb = CalamineWorkbook.from_path(file_path)
        return wb.sheet_names

    def import_data(
        self,
        conn: "DuckDBPyConnection",
        file_path: str,
        sheet: Optional[str] = None,
        header: bool = True,
    ) -> ImportResult:
        """Import Excel sheet into DuckDB.

        Reads data using openpyxl (.xlsx) or python-calamine (.xls)
        and inserts into DuckDB for SQL access.
        Empty rows are silently skipped per CONTEXT.md.

        Args:
            conn: DuckDB connection
            file_path: Path to Excel file (.xlsx or .xls)
            sheet: Sheet name to import (default: first sheet)
            header: Whether first row contains headers (default: True)

        Returns:
            ImportResult with schema and row count

        Raises:
            FileNotFoundError: If file does not exist
            ValueError: If file has no sheets
            ImportError: If .xls file and python-calamine not installed
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Get sheet name if not provided
        if sheet is None:
            sheets = self.list_sheets(file_path)
            if not sheets:
                raise ValueError("Excel file contains no sheets")
            sheet = sheets[0]

        # Route .xls to calamine, .xlsx to openpyxl
        if self._is_legacy_xls(file_path):
            rows_data = self._read_xls_calamine(file_path, sheet)
        else:
            rows_data = self._read_xlsx_openpyxl(file_path, sheet)

        if not rows_data:
            # Empty sheet - create empty table
            conn.execute("CREATE OR REPLACE TABLE imported_data (empty_sheet BOOLEAN)")
            return ImportResult(
                row_count=0,
                columns=[],
                warnings=["Sheet is empty"],
                source_type="excel",
            )

        # Extract headers
        if header:
            headers = [str(h) if h is not None else f"column_{i+1}"
                      for i, h in enumerate(rows_data[0])]
            data_rows = rows_data[1:]
        else:
            # Generate column names
            num_cols = len(rows_data[0]) if rows_data else 0
            headers = [f"column_{i+1}" for i in range(num_cols)]
            data_rows = rows_data

        # Ensure unique column names
        headers = self._make_unique_headers(headers)

        # Filter out empty rows (all None or empty string values)
        non_empty_rows = [
            row for row in data_rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        # Infer types from data
        column_types = self._infer_column_types(headers, non_empty_rows)

        # Create table with inferred schema + identity tracking column
        col_defs = f"{SOURCE_ROW_NUM_COLUMN} BIGINT, " + ", ".join([
            f'"{name}" {dtype}' for name, dtype in zip(headers, column_types)
        ])
        conn.execute(f"CREATE OR REPLACE TABLE imported_data ({col_defs})")

        # Insert data with 1-based source row numbers
        if non_empty_rows:
            placeholders = ", ".join(["?"] * (len(headers) + 1))
            for row_idx, row in enumerate(non_empty_rows, start=1):
                # Pad or truncate row to match header count
                padded_row = list(row)[:len(headers)]
                while len(padded_row) < len(headers):
                    padded_row.append(None)
                conn.execute(
                    f"INSERT INTO imported_data VALUES ({placeholders})",
                    [row_idx] + padded_row
                )

        # Build schema info with warnings
        columns = []
        all_warnings = []

        for name, dtype in zip(headers, column_types):
            col_warnings = []

            # Check for date columns and parse with warnings
            if "DATE" in dtype.upper() or "TIMESTAMP" in dtype.upper():
                sample = conn.execute(f"""
                    SELECT "{name}" FROM imported_data
                    WHERE "{name}" IS NOT NULL LIMIT 1
                """).fetchone()
                if sample:
                    date_result = parse_date_with_warnings(str(sample[0]))
                    for w in date_result.get("warnings", []):
                        msg = w.get("message", str(w)) if isinstance(w, dict) else str(w)
                        col_warnings.append(msg)

            # Check for nullable
            null_count = conn.execute(f"""
                SELECT COUNT(*) FROM imported_data WHERE "{name}" IS NULL
            """).fetchone()[0]
            is_nullable = null_count > 0

            columns.append(SchemaColumn(
                name=name,
                type=dtype,
                nullable=is_nullable,
                warnings=col_warnings,
            ))
            all_warnings.extend(col_warnings)

        # Get row count
        row_count = conn.execute("SELECT COUNT(*) FROM imported_data").fetchone()[0]

        # Track skipped rows
        skipped = len(data_rows) - len(non_empty_rows)
        if skipped > 0:
            all_warnings.append(f"Skipped {skipped} empty rows")

        return ImportResult(
            row_count=row_count,
            columns=columns,
            warnings=all_warnings,
            source_type="excel",
        )

    def _read_xlsx_openpyxl(
        self, file_path: str, sheet: str
    ) -> list[tuple]:
        """Read rows from .xlsx using openpyxl.

        Args:
            file_path: Path to .xlsx file.
            sheet: Sheet name to read.

        Returns:
            List of row tuples (including header row).
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet]
        rows_data = list(ws.values)
        wb.close()
        return rows_data

    def _read_xls_calamine(
        self, file_path: str, sheet: str
    ) -> list[tuple]:
        """Read rows from legacy .xls using python-calamine.

        Args:
            file_path: Path to .xls file.
            sheet: Sheet name to read.

        Returns:
            List of row tuples (including header row).

        Raises:
            ImportError: If python-calamine is not installed.
        """
        if not _calamine_available:
            raise ImportError(
                "python-calamine required for .xls files: "
                "pip install python-calamine"
            )
        wb = CalamineWorkbook.from_path(file_path)
        data = wb.get_sheet_by_name(sheet).to_python()
        # calamine returns list[list[Any]], convert to list[tuple] for compat
        return [tuple(row) for row in data]

    def get_metadata(self, conn: "DuckDBPyConnection") -> dict:
        """Get metadata about imported Excel data.

        Args:
            conn: Active DuckDB connection

        Returns:
            Dictionary with row_count, column_count, and source_type
        """
        try:
            row_count = conn.execute(
                "SELECT COUNT(*) FROM imported_data"
            ).fetchone()[0]
            columns = conn.execute("DESCRIBE imported_data").fetchall()
            # Exclude internal _source_row_num from user-facing count
            user_columns = [c for c in columns if c[0] != SOURCE_ROW_NUM_COLUMN]
            return {
                "row_count": row_count,
                "column_count": len(user_columns),
                "source_type": "excel",
            }
        except Exception:
            return {"error": "No data imported"}

    def _make_unique_headers(self, headers: list[str]) -> list[str]:
        """Ensure all headers are unique by appending suffixes.

        Args:
            headers: Original header names

        Returns:
            List of unique header names
        """
        seen: dict[str, int] = {}
        result = []

        for h in headers:
            name = h.strip() if h else "column"
            if not name:
                name = "column"

            if name in seen:
                seen[name] += 1
                result.append(f"{name}_{seen[name]}")
            else:
                seen[name] = 0
                result.append(name)

        return result

    def _infer_column_types(
        self, headers: list[str], rows: list[tuple]
    ) -> list[str]:
        """Infer DuckDB column types from sample data.

        Uses best-effort type inference with VARCHAR fallback.

        Args:
            headers: Column names
            rows: Data rows for type inference

        Returns:
            List of DuckDB type names
        """
        if not rows:
            return ["VARCHAR"] * len(headers)

        types = []
        for col_idx in range(len(headers)):
            # Collect non-null values for this column
            values = [
                row[col_idx] for row in rows
                if col_idx < len(row) and row[col_idx] is not None
            ]

            if not values:
                types.append("VARCHAR")
                continue

            # Check types of values
            sample_types = set()
            for v in values[:100]:  # Sample first 100 non-null values
                if isinstance(v, bool):
                    sample_types.add("BOOLEAN")
                elif isinstance(v, int):
                    sample_types.add("BIGINT")
                elif isinstance(v, float):
                    sample_types.add("DOUBLE")
                elif hasattr(v, 'isoformat'):  # datetime, date
                    sample_types.add("TIMESTAMP")
                else:
                    sample_types.add("VARCHAR")

            # Handle type selection
            if "VARCHAR" in sample_types:
                # Any string makes it VARCHAR
                types.append("VARCHAR")
            elif sample_types == {"BIGINT", "DOUBLE"} or sample_types == {"BIGINT"} | {"DOUBLE"}:
                # Mixed int/float is DOUBLE (integers can be stored as float)
                types.append("DOUBLE")
            elif len(sample_types) > 1:
                # Other mixed types become VARCHAR
                types.append("VARCHAR")
            elif "DOUBLE" in sample_types:
                types.append("DOUBLE")
            elif "BIGINT" in sample_types:
                types.append("BIGINT")
            elif "TIMESTAMP" in sample_types:
                types.append("TIMESTAMP")
            elif "BOOLEAN" in sample_types:
                types.append("BOOLEAN")
            else:
                types.append("VARCHAR")

        return types
