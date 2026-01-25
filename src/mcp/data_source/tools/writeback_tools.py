"""Write-back tools for Data Source MCP.

Provides MCP tools for writing tracking numbers and shipment data
back to the original data source after successful shipment creation.

Per CONTEXT.md Decision 4:
- Immediate write-back after each successful row
- Columns: tracking_number and shipped_at timestamp
- Atomic operations for file sources (temp + rename)
- shipped_at uses ISO8601 format
"""

import csv
import os
import tempfile
from datetime import datetime, timezone
from typing import Optional

from fastmcp import Context

# Import openpyxl lazily to avoid import error if not installed
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore


async def write_back(
    row_number: int,
    tracking_number: str,
    ctx: Context,
    shipped_at: Optional[str] = None,
) -> dict:
    """Write tracking number back to the original data source.

    Updates the original data source (CSV, Excel, or database) with
    the tracking number and shipment timestamp for a specific row.

    Args:
        row_number: 1-based row number to update
        tracking_number: UPS tracking number from shipment creation
        shipped_at: ISO8601 timestamp (default: current UTC time)

    Returns:
        Dictionary with:
        - success: True if write-back succeeded
        - source_type: Type of source updated (csv, excel, database)
        - row_number: Row that was updated
        - tracking_number: Tracking number written

    Raises:
        ValueError: If no source is loaded or source type is unsupported

    Example:
        >>> result = await write_back(1, "1Z999AA10123456784", ctx)
        >>> print(result["success"])
        True
    """
    # Get current source from lifespan context
    current_source = ctx.request_context.lifespan_context.get("current_source")

    if current_source is None:
        raise ValueError(
            "No data source loaded. Import a CSV, Excel, or database first."
        )

    # Default shipped_at to current UTC time
    if shipped_at is None:
        shipped_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    source_type = current_source.get("type")

    await ctx.info(
        f"Writing tracking number {tracking_number} to row {row_number} "
        f"in {source_type} source"
    )

    if source_type == "csv":
        await _write_back_csv(
            current_source["path"], row_number, tracking_number, shipped_at, ctx
        )
    elif source_type == "excel":
        await _write_back_excel(
            current_source["path"],
            row_number,
            tracking_number,
            shipped_at,
            ctx,
            sheet_name=current_source.get("sheet"),
        )
    elif source_type == "database":
        await _write_back_database(
            row_number, tracking_number, shipped_at, ctx
        )
    else:
        raise ValueError(f"Unsupported source type for write-back: {source_type}")

    await ctx.info(f"Successfully wrote tracking number to row {row_number}")

    return {
        "success": True,
        "source_type": source_type,
        "row_number": row_number,
        "tracking_number": tracking_number,
    }


async def _write_back_csv(
    file_path: str,
    row_number: int,
    tracking_number: str,
    shipped_at: str,
    ctx: Context,
) -> None:
    """Write tracking number to CSV file using atomic operations.

    Uses temp file + rename pattern for atomicity:
    1. Read original CSV
    2. Add/update tracking columns
    3. Write to temp file in same directory
    4. Atomic rename to original path

    Args:
        file_path: Path to CSV file
        row_number: 1-based row number to update
        tracking_number: Tracking number to write
        shipped_at: ISO8601 timestamp
        ctx: MCP context for logging
    """
    # Get directory for temp file (same dir for atomic rename)
    dir_path = os.path.dirname(file_path)
    temp_fd = None
    temp_path = None

    try:
        # Read original CSV
        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames or [])
            rows = list(reader)

        # Validate row number (row_number is 1-based, rows list is 0-based)
        if row_number < 1 or row_number > len(rows):
            raise ValueError(
                f"Row {row_number} not found. CSV has {len(rows)} data rows."
            )

        # Add tracking columns if not present
        if "tracking_number" not in fieldnames:
            fieldnames.append("tracking_number")
        if "shipped_at" not in fieldnames:
            fieldnames.append("shipped_at")

        # Update the target row (row_number is 1-based)
        target_row = rows[row_number - 1]
        target_row["tracking_number"] = tracking_number
        target_row["shipped_at"] = shipped_at

        # Write to temp file
        temp_fd, temp_path = tempfile.mkstemp(suffix=".csv", dir=dir_path)

        with os.fdopen(temp_fd, "w", newline="", encoding="utf-8") as f:
            temp_fd = None  # fdopen takes ownership
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        # Atomic rename
        os.replace(temp_path, file_path)
        temp_path = None  # Successfully replaced, don't clean up

        await ctx.info(f"CSV write-back complete for row {row_number}")

    except Exception:
        # Clean up temp file on error
        if temp_fd is not None:
            os.close(temp_fd)
        if temp_path is not None and os.path.exists(temp_path):
            os.unlink(temp_path)
        raise


async def _write_back_excel(
    file_path: str,
    row_number: int,
    tracking_number: str,
    shipped_at: str,
    ctx: Context,
    sheet_name: Optional[str] = None,
) -> None:
    """Write tracking number to Excel file using atomic operations.

    Uses temp file + rename pattern for atomicity:
    1. Load workbook
    2. Find or create tracking columns
    3. Update target row
    4. Save to temp file
    5. Atomic rename to original path

    Args:
        file_path: Path to Excel file
        row_number: 1-based row number to update (data row, not including header)
        tracking_number: Tracking number to write
        shipped_at: ISO8601 timestamp
        ctx: MCP context for logging
        sheet_name: Sheet name (optional, defaults to active sheet)
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required for Excel write-back")

    dir_path = os.path.dirname(file_path)
    temp_fd = None
    temp_path = None

    try:
        # Load workbook
        wb = load_workbook(file_path)

        # Get the target sheet
        if sheet_name and sheet_name != "(first sheet)":
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            raise ValueError("No active worksheet found")

        # Get header row (row 1)
        headers = [cell.value for cell in ws[1]]

        # Find or create tracking_number column
        tracking_col = None
        shipped_at_col = None

        for idx, header in enumerate(headers, start=1):
            if header == "tracking_number":
                tracking_col = idx
            elif header == "shipped_at":
                shipped_at_col = idx

        # Add tracking_number column if not present
        if tracking_col is None:
            tracking_col = len(headers) + 1
            ws.cell(row=1, column=tracking_col, value="tracking_number")

        # Add shipped_at column if not present
        if shipped_at_col is None:
            shipped_at_col = tracking_col + 1 if tracking_col == len(headers) + 1 else len(headers) + 2
            ws.cell(row=1, column=shipped_at_col, value="shipped_at")

        # Excel row = data row number + 1 (header is row 1)
        excel_row = row_number + 1

        # Validate row exists
        max_row = ws.max_row
        if excel_row > max_row or row_number < 1:
            raise ValueError(
                f"Row {row_number} not found. Excel has {max_row - 1} data rows."
            )

        # Update the target row
        ws.cell(row=excel_row, column=tracking_col, value=tracking_number)
        ws.cell(row=excel_row, column=shipped_at_col, value=shipped_at)

        # Write to temp file
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_path)
        os.close(temp_fd)
        temp_fd = None

        wb.save(temp_path)
        wb.close()

        # Atomic rename
        os.replace(temp_path, file_path)
        temp_path = None

        await ctx.info(f"Excel write-back complete for row {row_number}")

    except Exception:
        # Clean up temp file on error
        if temp_fd is not None:
            os.close(temp_fd)
        if temp_path is not None and os.path.exists(temp_path):
            os.unlink(temp_path)
        raise


async def _write_back_database(
    row_number: int,
    tracking_number: str,
    shipped_at: str,
    ctx: Context,
) -> None:
    """Write tracking number to database source.

    Uses DuckDB's attached database connection to execute UPDATE.
    Requires _row_number column to be present (added during import).

    Note: For MVP, this assumes the source table has been modified
    to include tracking_number and shipped_at columns, or the UPDATE
    will add them (depending on database support).

    Args:
        row_number: Row number to update (matches _row_number column)
        tracking_number: Tracking number to write
        shipped_at: ISO8601 timestamp
        ctx: MCP context for logging and database access
    """
    db = ctx.request_context.lifespan_context["db"]
    current_source = ctx.request_context.lifespan_context["current_source"]

    # For database sources, we need the table name from the query
    # This is a limitation - we only support simple table references
    # For complex queries, write-back would need to be handled differently
    query = current_source.get("query", "")

    # Extract table name from simple SELECT ... FROM table_name patterns
    # This is a simplification for MVP - complex queries may not work
    table_name = _extract_table_name(query)

    if table_name is None:
        raise ValueError(
            "Cannot determine target table for write-back. "
            "Database write-back requires a simple SELECT ... FROM table_name query."
        )

    await ctx.info(f"Updating database table {table_name} row {row_number}")

    # Use parameterized query to prevent SQL injection
    # Note: DuckDB uses $1, $2 syntax for parameters
    update_sql = f"""
        UPDATE {table_name}
        SET tracking_number = $1, shipped_at = $2
        WHERE _row_number = $3
    """

    try:
        db.execute(update_sql, [tracking_number, shipped_at, row_number])
        await ctx.info(f"Database write-back complete for row {row_number}")
    except Exception as e:
        raise ValueError(f"Database write-back failed: {e}") from e


def _extract_table_name(query: str) -> Optional[str]:
    """Extract table name from a simple SELECT query.

    Handles patterns like:
    - SELECT * FROM table_name
    - SELECT * FROM table_name WHERE ...
    - SELECT cols FROM schema.table_name

    Args:
        query: SQL query string

    Returns:
        Table name or None if not extractable
    """
    if not query:
        return None

    # Normalize whitespace
    query = " ".join(query.split())
    query_upper = query.upper()

    # Find FROM keyword
    from_idx = query_upper.find(" FROM ")
    if from_idx == -1:
        return None

    # Extract portion after FROM
    after_from = query[from_idx + 6:].strip()

    # Get the table name (first token after FROM)
    # Handle schema.table notation
    parts = after_from.split()
    if not parts:
        return None

    table_ref = parts[0]

    # Check if the table reference itself IS a keyword (shouldn't be in valid SQL)
    # Don't check for substring match as "orders" contains "ORDER"
    keywords = {"WHERE", "ORDER", "GROUP", "HAVING", "LIMIT", "OFFSET", "JOIN", "LEFT", "RIGHT", "INNER", "OUTER"}
    if table_ref.upper() in keywords:
        return None

    # Reject subqueries - table reference shouldn't start with parenthesis
    if table_ref.startswith("("):
        return None

    return table_ref
