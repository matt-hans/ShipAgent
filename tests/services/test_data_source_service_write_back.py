"""Tests for DataSourceService batch write-back and replay recovery paths."""

import csv
import json
import os
import tempfile
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.db.models import AuditLog, Base, Job, JobRow, RowStatus
from src.services.data_source_service import (
    SOURCE_SIGNATURE_AUDIT_MESSAGE,
    DataSourceService,
)


@pytest.fixture(autouse=True)
def reset_data_source_service():
    """Ensure tests don't leak singleton state across cases."""
    DataSourceService.reset_instance()
    yield
    DataSourceService.reset_instance()


@pytest.fixture
def sample_csv_file():
    """Create a temporary CSV with three data rows."""
    content = """order_id,customer_name,state
1001,John Doe,CA
1002,Jane Smith,NY
1003,Bob Wilson,TX
"""
    fd, path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        yield path
    finally:
        Path(path).unlink(missing_ok=True)


@pytest.fixture
def sample_excel_file():
    """Create a temporary XLSX with three data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["order_id", "customer_name", "state"])
    ws.append([1001, "John Doe", "CA"])
    ws.append([1002, "Jane Smith", "NY"])
    ws.append([1003, "Bob Wilson", "TX"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name
    Path(f.name).unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_write_back_batch_csv_success(sample_csv_file):
    """CSV batch write-back updates all requested rows in one call."""
    svc = DataSourceService.get_instance()
    await svc.import_csv(sample_csv_file)

    result = await svc.write_back_batch(
        [
            (1, "1Z999AA10123456784"),
            (3, "1Z999AA10123456786"),
        ]
    )

    assert result["status"] == "success"
    assert result["written"] == 2

    with open(sample_csv_file, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    assert rows[0]["tracking_number"] == "1Z999AA10123456784"
    assert rows[1]["tracking_number"] == ""
    assert rows[2]["tracking_number"] == "1Z999AA10123456786"


@pytest.mark.asyncio
async def test_write_back_batch_excel_success(sample_excel_file):
    """Excel batch write-back has parity with CSV batch behavior."""
    svc = DataSourceService.get_instance()
    await svc.import_excel(sample_excel_file, sheet="Orders")

    result = await svc.write_back_batch(
        [
            (1, "1Z999AA10123456784"),
            (2, "1Z999AA10123456785"),
        ]
    )

    assert result["status"] == "success"
    assert result["written"] == 2

    wb = load_workbook(sample_excel_file)
    ws = wb["Orders"]
    headers = [cell.value for cell in ws[1]]
    tracking_col = headers.index("tracking_number") + 1
    assert ws.cell(row=2, column=tracking_col).value == "1Z999AA10123456784"
    assert ws.cell(row=3, column=tracking_col).value == "1Z999AA10123456785"
    wb.close()


@pytest.mark.asyncio
async def test_write_back_batch_csv_is_atomic_when_replace_fails(sample_csv_file):
    """Original CSV remains untouched when atomic replace fails."""
    svc = DataSourceService.get_instance()
    await svc.import_csv(sample_csv_file)

    with open(sample_csv_file, "r", encoding="utf-8") as f:
        original_content = f.read()

    with patch(
        "src.services.write_back_utils.os.replace",
        side_effect=OSError("replace failed"),
    ):
        result = await svc.write_back_batch([(1, "1Z999AA10123456784")])

    assert result["status"] == "error"

    with open(sample_csv_file, "r", encoding="utf-8") as f:
        assert f.read() == original_content

    temp_files = [
        name
        for name in os.listdir(os.path.dirname(sample_csv_file))
        if name.endswith(".csv.tmp")
    ]
    assert temp_files == []


@pytest.mark.asyncio
async def test_replay_write_back_from_job_uses_persisted_source_signature(
    sample_csv_file,
):
    """Replay reapplies tracking numbers from JobRow when source signature matches."""
    svc = DataSourceService.get_instance()
    await svc.import_csv(sample_csv_file)
    signature = svc.get_source_signature()
    assert signature is not None

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    try:
        job = Job(
            id="job-replay-1",
            name="Replay Job",
            original_command="Ship all orders",
            status="failed",
            total_rows=2,
        )
        db.add(job)
        db.add(
            JobRow(
                job_id=job.id,
                row_number=1,
                row_checksum="checksum-1",
                status=RowStatus.completed.value,
                tracking_number="1Z999AA10123456784",
            ),
        )
        db.add(
            JobRow(
                job_id=job.id,
                row_number=2,
                row_checksum="checksum-2",
                status=RowStatus.failed.value,
                tracking_number=None,
            ),
        )
        db.add(
            AuditLog(
                job_id=job.id,
                level="INFO",
                event_type="row_event",
                message=SOURCE_SIGNATURE_AUDIT_MESSAGE,
                details=json.dumps({"source_signature": signature}),
            ),
        )
        db.commit()

        @contextmanager
        def _fake_db_context():
            yield db

        with patch("src.db.connection.get_db_context", _fake_db_context):
            replay_result = await svc.replay_write_back_from_job(job.id)
    finally:
        db.close()

    assert replay_result["status"] == "success"
    assert replay_result["replayed_rows"] == 1

    with open(sample_csv_file, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["tracking_number"] == "1Z999AA10123456784"


@pytest.mark.asyncio
async def test_replay_write_back_fails_when_source_signature_mismatches(
    sample_csv_file,
):
    """Replay hard-fails when the currently loaded source is not the original source."""
    svc = DataSourceService.get_instance()
    await svc.import_csv(sample_csv_file)
    signature = svc.get_source_signature()
    assert signature is not None

    with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False) as f:
        f.write("order_id,customer_name,state\n9999,Other,WA\n")
        other_csv = f.name

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    try:
        job = Job(
            id="job-replay-2",
            name="Replay Job",
            original_command="Ship all orders",
            status="failed",
            total_rows=1,
        )
        db.add(job)
        db.add(
            JobRow(
                job_id=job.id,
                row_number=1,
                row_checksum="checksum-1",
                status=RowStatus.completed.value,
                tracking_number="1Z999AA10123456784",
            ),
        )
        db.add(
            AuditLog(
                job_id=job.id,
                level="INFO",
                event_type="row_event",
                message=SOURCE_SIGNATURE_AUDIT_MESSAGE,
                details=json.dumps({"source_signature": signature}),
            ),
        )
        db.commit()

        await svc.import_csv(other_csv)

        @contextmanager
        def _fake_db_context():
            yield db

        with patch("src.db.connection.get_db_context", _fake_db_context):
            replay_result = await svc.replay_write_back_from_job(job.id)
    finally:
        db.close()
        Path(other_csv).unlink(missing_ok=True)

    assert replay_result["status"] == "error"
    assert "does not match job source" in replay_result["message"]
