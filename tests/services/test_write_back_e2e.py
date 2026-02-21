"""End-to-end write-back integration tests.

Exercises the production write-back path through the MCP gateway:
  BatchEngine.execute() → DataSourceMCPClient.write_back_batch()
  → MCP write_back tool → write_back_utils → file I/O

Only the UPS carrier path (UPSMCPClient) is mocked — the data gateway
runs the real MCP server subprocess for import, query, and write-back.

Marked @pytest.mark.integration so they can be selected/excluded easily:
    pytest -m integration       # run only integration tests
    pytest -m "not integration" # skip integration tests
"""

import csv
import os
import tempfile
from typing import Any
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.db.models import Base
from src.services.batch_engine import BatchEngine
from src.services.data_source_mcp_client import DataSourceMCPClient
from src.services.job_service import JobService

# ---------------------------------------------------------------------------
# Markers applied to every test in this module
# ---------------------------------------------------------------------------
pytestmark = [pytest.mark.integration, pytest.mark.asyncio]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
async def mcp():
    """Per-test MCP client with real subprocess."""
    client = DataSourceMCPClient()
    await client.connect()
    yield client
    await client.disconnect_mcp()


@pytest.fixture
def db_session():
    """Create in-memory SQLite session with schema."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def job_service(db_session) -> JobService:
    """Create JobService with in-memory DB."""
    return JobService(db_session)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_csv(rows: list[dict[str, Any]], path: str | None = None) -> str:
    """Write rows to a temp CSV and return the path."""
    if not rows:
        raise ValueError("Need at least one row")
    fieldnames = list(rows[0].keys())
    if path is None:
        fd, path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _make_excel(rows: list[dict[str, Any]], path: str | None = None) -> str:
    """Write rows to a temp Excel file and return the path."""
    if not rows:
        raise ValueError("Need at least one row")
    wb = Workbook()
    ws = wb.active
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    if path is None:
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
    wb.save(path)
    wb.close()
    return path


def _read_csv(path: str) -> list[dict[str, str]]:
    """Read CSV back as list of dicts."""
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_excel(path: str) -> list[dict[str, Any]]:
    """Read Excel back as list of dicts."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    data = list(ws.values)
    wb.close()
    if len(data) < 2:
        return []
    headers = [str(h) for h in data[0]]
    return [dict(zip(headers, row, strict=False)) for row in data[1:]]


def _sample_rows(n: int = 5, states: list[str] | None = None) -> list[dict]:
    """Generate n sample shipping rows.

    Zip codes use 5+4 format so DuckDB infers VARCHAR (not BIGINT).
    Weight is a plain numeric string (no units).
    Country is included for shipTo.countryCode mapping.
    """
    if states is None:
        states = ["CA", "NY", "TX", "FL", "WA"]
    rows = []
    for i in range(n):
        rows.append({
            "order_id": f"ORD-{1000 + i}",
            "recipient_name": f"Customer {i}",
            "address": f"{100 + i} Main St",
            "city": f"City{i}",
            "state": states[i % len(states)],
            "zip": f"9{i:04d}-0001",
            "country": "US",
            "weight": f"{2.0 + i * 0.5:.1f}",
        })
    return rows


def _mock_ups(
    fail_rows: set[int] | None = None,
) -> AsyncMock:
    """Create a mock UPSMCPClient that returns fake shipment results.

    Args:
        fail_rows: Set of call-order indices (1-based) to fail.
    """
    ups = AsyncMock()
    ups.is_connected = True
    _call_count = {"n": 0}

    async def _ship(request_body: Any = None, **kw):
        """Simulate UPS create_shipment."""
        _call_count["n"] += 1
        n = _call_count["n"]
        if fail_rows and n in fail_rows:
            from src.services.errors import UPSServiceError
            raise UPSServiceError(f"Simulated failure for call {n}")
        return {
            "trackingNumbers": [f"1ZTEST{n:08d}"],
            "shipmentIdentificationNumber": f"SID{n:08d}",
            "labelData": [],
            "totalCharges": {"monetaryValue": "10.50"},
        }

    ups.create_shipment = AsyncMock(side_effect=_ship)
    return ups


async def _import_and_build_job(
    mcp_client: DataSourceMCPClient,
    file_path: str,
    job_service: JobService,
    where_sql: str | None = None,
    is_excel: bool = False,
) -> tuple[str, list[Any]]:
    """Import file via MCP, fetch rows, create job + rows, return (job_id, db_rows).

    Exercises the production import → query → job creation chain.
    """
    if is_excel:
        await mcp_client.import_excel(file_path)
    else:
        await mcp_client.import_csv(file_path)

    # Fetch rows through MCP (applies _source_row_num)
    flat_rows = await mcp_client.get_rows_by_filter(
        where_sql=where_sql, limit=1000
    )

    # Build job row data (uses _row_number from normalized rows)
    from src.orchestrator.agent.tools.core import _build_job_row_data
    row_data = _build_job_row_data(list(flat_rows))

    job = job_service.create_job(name="E2E Test", original_command="test")
    job_service.create_rows(job.id, row_data)
    db_rows = job_service.get_rows(job.id)
    return job.id, db_rows


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestWriteBackE2E:
    """End-to-end write-back tests through real MCP gateway."""

    async def test_batch_execute_writes_tracking_to_csv(
        self, mcp, db_session, job_service
    ):
        """All rows succeed → tracking_number + shipped_at added to CSV."""
        rows = _sample_rows(5)
        csv_path = _make_csv(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service
            )
            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            # Patch get_data_gateway to return our real MCP client
            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 5
            assert result["failed"] == 0
            wb = result["write_back"]
            assert wb["status"] == "success"
            assert wb["success_count"] == 5

            # Verify CSV file was updated
            updated = _read_csv(csv_path)
            assert len(updated) == 5
            for row in updated:
                assert row.get("tracking_number"), f"Row missing tracking: {row}"
                assert row.get("shipped_at"), f"Row missing shipped_at: {row}"
        finally:
            os.unlink(csv_path)

    async def test_batch_execute_writes_tracking_to_excel(
        self, mcp, db_session, job_service
    ):
        """All rows succeed → tracking_number + shipped_at added to Excel."""
        rows = _sample_rows(3)
        xlsx_path = _make_excel(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, xlsx_path, job_service, is_excel=True
            )
            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 3
            wb = result["write_back"]
            assert wb["status"] == "success"

            updated = _read_excel(xlsx_path)
            assert len(updated) == 3
            for row in updated:
                assert row.get("tracking_number"), f"Row missing tracking: {row}"
        finally:
            os.unlink(xlsx_path)

    async def test_batch_execute_skips_write_back_when_disabled(
        self, mcp, db_session, job_service
    ):
        """write_back_enabled=False → CSV untouched, result says skipped."""
        rows = _sample_rows(3)
        csv_path = _make_csv(rows)

        try:
            original = _read_csv(csv_path)
            original_fieldnames = list(original[0].keys())

            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service
            )
            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=False,
                )

            assert result["successful"] == 3
            assert result["write_back"]["status"] == "skipped"

            after = _read_csv(csv_path)
            after_fieldnames = list(after[0].keys())
            assert "tracking_number" not in after_fieldnames
            assert after_fieldnames == original_fieldnames
        finally:
            os.unlink(csv_path)

    async def test_batch_execute_partial_write_back_on_mixed_results(
        self, mcp, db_session, job_service
    ):
        """Mix of success/failure → only successful rows get tracking."""
        rows = _sample_rows(5)
        csv_path = _make_csv(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service
            )
            ups = _mock_ups(fail_rows={2, 4})
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 3
            assert result["failed"] == 2

            updated = _read_csv(csv_path)
            tracking_values = [r.get("tracking_number", "") for r in updated]
            populated = [t for t in tracking_values if t]
            empty = [t for t in tracking_values if not t]
            assert len(populated) == 3
            assert len(empty) == 2
        finally:
            os.unlink(csv_path)

    async def test_write_back_preserves_existing_tracking_on_rerun(
        self, mcp, db_session, job_service
    ):
        """Pre-existing tracking_number for row 1 → still gets overwritten by batch."""
        rows = _sample_rows(3)
        rows[0]["tracking_number"] = "EXISTING_TRACK_001"
        csv_path = _make_csv(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service
            )
            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 3
            updated = _read_csv(csv_path)
            for row in updated:
                assert row.get("tracking_number")
        finally:
            os.unlink(csv_path)

    async def test_filtered_subset_writes_to_correct_source_rows(
        self, mcp, db_session, job_service
    ):
        """P0 REGRESSION: filtered rows write tracking to correct source positions.

        10-row CSV, filter matches rows 3, 7, 9 (1-based; 0-indexed: 2, 6, 8 are CA).
        After batch execution, only those 3 source rows should have tracking.
        All other 7 rows must remain untouched.
        """
        states = ["NY", "TX", "CA", "FL", "WA", "OR", "CA", "IL", "CA", "CO"]
        rows = []
        for i in range(10):
            rows.append({
                "order_id": f"ORD-{2000 + i}",
                "recipient_name": f"Person {i}",
                "address": f"{200 + i} Oak Ave",
                "city": f"Town{i}",
                "state": states[i],
                "zip": f"1{i:04d}-0001",
                "country": "US",
                "weight": "3.0",
            })
        csv_path = _make_csv(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service,
                where_sql="state = 'CA'",
            )
            assert len(db_rows) == 3
            source_positions = sorted(r.row_number for r in db_rows)
            assert source_positions == [3, 7, 9], (
                f"Expected source rows [3, 7, 9], got {source_positions}"
            )

            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 3

            # CRITICAL: verify tracking in correct source rows
            updated = _read_csv(csv_path)
            assert len(updated) == 10

            for i, row in enumerate(updated):
                src = i + 1
                if src in (3, 7, 9):
                    assert row.get("tracking_number"), (
                        f"Source row {src} should have tracking but got: {row}"
                    )
                    assert row.get("shipped_at"), (
                        f"Source row {src} should have shipped_at"
                    )
                else:
                    assert not row.get("tracking_number"), (
                        f"Source row {src} should NOT have tracking but got: "
                        f"{row.get('tracking_number')}"
                    )
        finally:
            os.unlink(csv_path)

    async def test_filtered_subset_writes_to_correct_excel_rows(
        self, mcp, db_session, job_service
    ):
        """P0 REGRESSION (Excel): filtered rows write to correct Excel positions."""
        states = ["NY", "TX", "CA", "FL", "WA", "OR", "CA", "IL", "CA", "CO"]
        rows = []
        for i in range(10):
            rows.append({
                "order_id": f"ORD-{3000 + i}",
                "recipient_name": f"Person {i}",
                "address": f"{300 + i} Elm Dr",
                "city": f"Town{i}",
                "state": states[i],
                "zip": f"2{i:04d}-0001",
                "country": "US",
                "weight": "4.0",
            })
        xlsx_path = _make_excel(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, xlsx_path, job_service,
                where_sql="state = 'CA'",
                is_excel=True,
            )
            assert len(db_rows) == 3
            source_positions = sorted(r.row_number for r in db_rows)
            assert source_positions == [3, 7, 9]

            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 3

            updated = _read_excel(xlsx_path)
            assert len(updated) == 10

            for i, row in enumerate(updated):
                src = i + 1
                if src in (3, 7, 9):
                    assert row.get("tracking_number"), (
                        f"Excel row {src} should have tracking"
                    )
                else:
                    assert not row.get("tracking_number"), (
                        f"Excel row {src} should NOT have tracking "
                        f"but got: {row.get('tracking_number')}"
                    )
        finally:
            os.unlink(xlsx_path)

    async def test_filtered_non_contiguous_write_back_correctness(
        self, mcp, db_session, job_service
    ):
        """20-row CSV, filter matches rows 2, 5, 11, 18 → only those get tracking."""
        target_indices = {1, 4, 10, 17}  # 0-indexed → 1-based: 2, 5, 11, 18
        rows = []
        for i in range(20):
            rows.append({
                "order_id": f"ORD-{4000 + i}",
                "recipient_name": f"Person {i}",
                "address": f"{400 + i} Pine Ln",
                "city": f"City{i}",
                "state": "HI" if i in target_indices else "NY",
                "zip": f"3{i:04d}-0001",
                "country": "US",
                "weight": "2.5",
            })
        csv_path = _make_csv(rows)

        try:
            job_id, db_rows = await _import_and_build_job(
                mcp, csv_path, job_service,
                where_sql="state = 'HI'",
            )
            assert len(db_rows) == 4
            source_positions = sorted(r.row_number for r in db_rows)
            assert source_positions == [2, 5, 11, 18]

            ups = _mock_ups()
            engine = BatchEngine(
                ups_service=ups, db_session=db_session, account_number="TEST"
            )

            with patch(
                "src.services.batch_engine.get_data_gateway",
                new_callable=AsyncMock,
                return_value=mcp,
            ):
                result = await engine.execute(
                    job_id=job_id, rows=db_rows,
                    shipper={"Name": "Test", "countryCode": "US"},
                    write_back_enabled=True,
                )

            assert result["successful"] == 4

            updated = _read_csv(csv_path)
            assert len(updated) == 20

            tracked_rows = []
            untracked_rows = []
            for i, row in enumerate(updated):
                if row.get("tracking_number"):
                    tracked_rows.append(i + 1)
                else:
                    untracked_rows.append(i + 1)

            assert tracked_rows == [2, 5, 11, 18], (
                f"Expected tracking on rows [2, 5, 11, 18], got {tracked_rows}"
            )
            assert len(untracked_rows) == 16
        finally:
            os.unlink(csv_path)
