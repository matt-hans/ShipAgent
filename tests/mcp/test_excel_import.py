"""Test Excel import functionality.

Tests the ExcelAdapter class for importing Excel files (.xlsx) into DuckDB.
Verifies sheet listing, default sheet import, specific sheet selection,
and error handling.
"""

import tempfile
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook

from src.mcp.data_source.adapters.excel_adapter import ExcelAdapter
from src.mcp.data_source.models import ImportResult


@pytest.fixture
def sample_excel():
    """Create a temporary Excel file for testing.

    Creates workbook with two sheets:
    - "Orders": order_id, customer, amount, state (3 rows)
    - "Returns": return_id, order_id, reason (2 rows)
    """
    wb = Workbook()

    # First sheet (default)
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["order_id", "customer", "amount", "state"])
    ws1.append([1001, "John Doe", 150.50, "CA"])
    ws1.append([1002, "Jane Smith", 200.00, "NY"])
    ws1.append([1003, "Bob Wilson", 75.25, "TX"])

    # Second sheet
    ws2 = wb.create_sheet("Returns")
    ws2.append(["return_id", "order_id", "reason"])
    ws2.append([1, 1001, "Damaged"])
    ws2.append([2, 1002, "Wrong item"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name

    # Cleanup
    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def sample_excel_with_empty_rows():
    """Create Excel file with empty rows to test skip behavior."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([None, None])  # Empty row
    ws.append([2, "Bob"])
    ws.append(["", ""])  # Empty string row
    ws.append([3, "Charlie"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name

    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def duckdb_conn():
    """Create in-memory DuckDB connection."""
    conn = duckdb.connect(":memory:")
    yield conn
    conn.close()


class TestListSheets:
    """Tests for ExcelAdapter.list_sheets method."""

    def test_list_sheets_returns_all_sheets(self, sample_excel):
        """Test that list_sheets returns all sheet names."""
        adapter = ExcelAdapter()
        sheets = adapter.list_sheets(sample_excel)

        assert len(sheets) == 2
        assert "Orders" in sheets
        assert "Returns" in sheets

    def test_list_sheets_preserves_order(self, sample_excel):
        """Test that sheets are returned in workbook order."""
        adapter = ExcelAdapter()
        sheets = adapter.list_sheets(sample_excel)

        # Orders is the first sheet, Returns is second
        assert sheets[0] == "Orders"
        assert sheets[1] == "Returns"

    def test_list_sheets_file_not_found(self):
        """Test error handling for missing file."""
        adapter = ExcelAdapter()
        with pytest.raises(FileNotFoundError):
            adapter.list_sheets("/nonexistent/file.xlsx")


class TestExcelImport:
    """Tests for ExcelAdapter.import_data method."""

    def test_import_default_sheet(self, sample_excel, duckdb_conn):
        """Test importing first sheet by default."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel)

        assert isinstance(result, ImportResult)
        assert result.row_count == 3
        assert result.source_type == "excel"

        # Should have imported Orders sheet
        col_names = [c.name.lower() for c in result.columns]
        assert "order_id" in col_names
        assert "customer" in col_names
        assert "amount" in col_names
        assert "state" in col_names

    def test_import_specific_sheet(self, sample_excel, duckdb_conn):
        """Test importing a specific sheet."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel, sheet="Returns")

        assert result.row_count == 2
        col_names = [c.name.lower() for c in result.columns]
        assert "return_id" in col_names
        assert "order_id" in col_names
        assert "reason" in col_names

    def test_import_creates_table(self, sample_excel, duckdb_conn):
        """Test that import creates queryable table."""
        adapter = ExcelAdapter()
        adapter.import_data(duckdb_conn, sample_excel)

        # Should be able to query the table
        rows = duckdb_conn.execute("SELECT * FROM imported_data").fetchall()
        assert len(rows) == 3

    def test_import_replaces_previous(self, sample_excel, duckdb_conn):
        """Test that importing replaces previous data."""
        adapter = ExcelAdapter()

        # Import Orders (3 rows)
        result1 = adapter.import_data(duckdb_conn, sample_excel, sheet="Orders")
        assert result1.row_count == 3

        # Import Returns (2 rows)
        result2 = adapter.import_data(duckdb_conn, sample_excel, sheet="Returns")
        assert result2.row_count == 2

        # Table should have 2 rows now (replaced)
        count = duckdb_conn.execute(
            "SELECT COUNT(*) FROM imported_data"
        ).fetchone()[0]
        assert count == 2

    def test_import_file_not_found(self, duckdb_conn):
        """Test error handling for missing file."""
        adapter = ExcelAdapter()
        with pytest.raises(FileNotFoundError):
            adapter.import_data(duckdb_conn, "/nonexistent/file.xlsx")


class TestEmptyRowHandling:
    """Tests for empty row skip behavior per CONTEXT.md."""

    def test_empty_rows_skipped(self, sample_excel_with_empty_rows, duckdb_conn):
        """Test that empty rows are silently skipped."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel_with_empty_rows)

        # Should have 3 data rows (id=1,2,3), skipping empty rows
        assert result.row_count == 3

        # Verify the actual data
        rows = duckdb_conn.execute(
            "SELECT id FROM imported_data ORDER BY id"
        ).fetchall()
        assert [r[0] for r in rows] == [1, 2, 3]

    def test_empty_rows_warning(self, sample_excel_with_empty_rows, duckdb_conn):
        """Test that skipped rows generate a warning."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel_with_empty_rows)

        # Should have warning about skipped rows
        assert any("Skipped" in w and "empty" in w.lower() for w in result.warnings)


class TestTypeInference:
    """Tests for type inference during import."""

    def test_numeric_types_inferred(self, sample_excel, duckdb_conn):
        """Test that numeric columns get appropriate types."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel)

        # Find amount column type
        amount_col = next(
            (c for c in result.columns if c.name.lower() == "amount"), None
        )
        assert amount_col is not None
        # Should be DOUBLE for float values
        assert "DOUBLE" in amount_col.type.upper() or "FLOAT" in amount_col.type.upper()

    def test_integer_types_inferred(self, sample_excel, duckdb_conn):
        """Test that integer columns get appropriate types."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel)

        # Find order_id column type
        order_id_col = next(
            (c for c in result.columns if c.name.lower() == "order_id"), None
        )
        assert order_id_col is not None
        # Should be BIGINT for integer values
        assert "INT" in order_id_col.type.upper()

    def test_string_types_inferred(self, sample_excel, duckdb_conn):
        """Test that string columns get VARCHAR type."""
        adapter = ExcelAdapter()
        result = adapter.import_data(duckdb_conn, sample_excel)

        # Find customer column type
        customer_col = next(
            (c for c in result.columns if c.name.lower() == "customer"), None
        )
        assert customer_col is not None
        assert "VARCHAR" in customer_col.type.upper()


class TestSourceType:
    """Tests for source_type property."""

    def test_source_type_is_excel(self):
        """Test that source_type returns 'excel'."""
        adapter = ExcelAdapter()
        assert adapter.source_type == "excel"


class TestGetMetadata:
    """Tests for ExcelAdapter.get_metadata method."""

    def test_get_metadata_after_import(self, sample_excel, duckdb_conn):
        """Test metadata after successful import."""
        adapter = ExcelAdapter()
        adapter.import_data(duckdb_conn, sample_excel)

        metadata = adapter.get_metadata(duckdb_conn)

        assert metadata["row_count"] == 3
        assert metadata["column_count"] == 4
        assert metadata["source_type"] == "excel"

    def test_get_metadata_no_data(self, duckdb_conn):
        """Test metadata when no data imported."""
        adapter = ExcelAdapter()
        metadata = adapter.get_metadata(duckdb_conn)

        assert "error" in metadata
