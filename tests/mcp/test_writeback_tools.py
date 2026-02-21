"""Test write_back tool functionality.

Tests the write_back MCP tool for persisting tracking numbers
back to CSV, Excel, and database sources.

Tests verify:
- Atomic file operations (temp + rename pattern)
- Column creation for tracking_number and shipped_at
- Row number correctness
- Error handling for edge cases
"""

import csv
import os
import tempfile
from datetime import UTC, datetime, timedelta
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import duckdb
import pytest
from openpyxl import Workbook, load_workbook

from src.mcp.data_source.models import SOURCE_ROW_NUM_COLUMN
from src.mcp.data_source.tools.writeback_tools import (
    _extract_table_name,
    write_back,
)


@pytest.fixture
def sample_csv():
    """Create a temporary CSV file for testing.

    Contains 3 data rows with order, customer, state.
    No tracking columns initially.
    """
    content = """order_id,customer_name,state
1001,John Doe,CA
1002,Jane Smith,NY
1003,Bob Wilson,TX"""

    # Create temp file, write content, and close it before yielding
    # This ensures the file is fully written and flushed
    fd, path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        yield path
    finally:
        Path(path).unlink(missing_ok=True)


@pytest.fixture
def sample_csv_with_tracking():
    """Create CSV that already has tracking_number column."""
    content = """order_id,customer_name,state,tracking_number,shipped_at
1001,John Doe,CA,,
1002,Jane Smith,NY,,
1003,Bob Wilson,TX,,"""

    # Create temp file, write content, and close it before yielding
    fd, path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        yield path
    finally:
        Path(path).unlink(missing_ok=True)


@pytest.fixture
def sample_excel():
    """Create a temporary Excel file for testing.

    Contains 3 data rows on Sheet1.
    No tracking columns initially.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["order_id", "customer_name", "state"])
    ws.append([1001, "John Doe", "CA"])
    ws.append([1002, "Jane Smith", "NY"])
    ws.append([1003, "Bob Wilson", "TX"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name

    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def sample_excel_with_tracking():
    """Create Excel that already has tracking columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["order_id", "customer_name", "state", "tracking_number", "shipped_at"])
    ws.append([1001, "John Doe", "CA", None, None])
    ws.append([1002, "Jane Smith", "NY", None, None])
    ws.append([1003, "Bob Wilson", "TX", None, None])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        yield f.name

    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def mock_ctx():
    """Create mock FastMCP context with lifespan_context."""
    ctx = AsyncMock()
    ctx.request_context = MagicMock()
    ctx.request_context.lifespan_context = {
        "db": MagicMock(),
        "current_source": None,
        "type_overrides": {},
    }
    return ctx


@pytest.fixture
def duckdb_conn():
    """Create in-memory DuckDB connection."""
    conn = duckdb.connect(":memory:")
    yield conn
    conn.close()


class TestWriteBackCSV:
    """Tests for CSV write-back functionality."""

    @pytest.mark.asyncio
    async def test_write_back_csv_adds_columns(self, sample_csv, mock_ctx):
        """Test that tracking columns are added when missing."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv,
            "row_count": 3,
        }

        result = await write_back(
            row_number=1,
            tracking_number="1Z999AA10123456784",
            ctx=mock_ctx,
            shipped_at="2026-01-25T12:00:00Z",
        )

        assert result["success"] is True
        assert result["source_type"] == "csv"
        assert result["row_number"] == 1
        assert result["tracking_number"] == "1Z999AA10123456784"

        # Verify columns were added
        with open(sample_csv) as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            assert "tracking_number" in fieldnames
            assert "shipped_at" in fieldnames

            # Verify correct row was updated
            rows = list(reader)
            assert rows[0]["tracking_number"] == "1Z999AA10123456784"
            assert rows[0]["shipped_at"] == "2026-01-25T12:00:00Z"

    @pytest.mark.asyncio
    async def test_write_back_csv_updates_existing(
        self, sample_csv_with_tracking, mock_ctx
    ):
        """Test updating existing tracking_number column."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv_with_tracking,
            "row_count": 3,
        }

        result = await write_back(
            row_number=2,
            tracking_number="1Z999AA10123456785",
            ctx=mock_ctx,
            shipped_at="2026-01-25T14:00:00Z",
        )

        assert result["success"] is True

        # Verify row 2 was updated
        with open(sample_csv_with_tracking) as f:
            rows = list(csv.DictReader(f))
            assert rows[1]["tracking_number"] == "1Z999AA10123456785"
            assert rows[1]["shipped_at"] == "2026-01-25T14:00:00Z"
            # Row 1 and 3 should be unchanged
            assert rows[0]["tracking_number"] == ""
            assert rows[2]["tracking_number"] == ""

    @pytest.mark.asyncio
    async def test_write_back_csv_preserves_data(self, sample_csv, mock_ctx):
        """Test that other columns remain unchanged."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv,
            "row_count": 3,
        }

        await write_back(
            row_number=1,
            tracking_number="1Z999AA10123456784",
            ctx=mock_ctx,
        )

        # Verify original data preserved
        with open(sample_csv) as f:
            rows = list(csv.DictReader(f))
            assert rows[0]["order_id"] == "1001"
            assert rows[0]["customer_name"] == "John Doe"
            assert rows[0]["state"] == "CA"
            # Other rows unchanged
            assert rows[1]["order_id"] == "1002"
            assert rows[2]["order_id"] == "1003"

    @pytest.mark.asyncio
    async def test_write_back_csv_default_shipped_at(self, sample_csv, mock_ctx):
        """Test that shipped_at defaults to current UTC time."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv,
            "row_count": 3,
        }

        # Truncate to seconds since shipped_at format doesn't include microseconds
        before = datetime.now(UTC).replace(microsecond=0)
        await write_back(
            row_number=1,
            tracking_number="1Z999AA10123456784",
            ctx=mock_ctx,
        )
        # Add 1 second buffer for timing margin
        after = datetime.now(UTC).replace(microsecond=0) + timedelta(seconds=1)

        with open(sample_csv) as f:
            rows = list(csv.DictReader(f))
            shipped_at = datetime.fromisoformat(
                rows[0]["shipped_at"].replace("Z", "+00:00")
            )
            assert before <= shipped_at <= after

    @pytest.mark.asyncio
    async def test_write_back_csv_atomic_on_error(self, sample_csv, mock_ctx):
        """Test that temp file is cleaned up on error."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv,
            "row_count": 3,
        }

        # Read original content
        with open(sample_csv) as f:
            original_content = f.read()

        # Try to update non-existent row
        with pytest.raises(ValueError, match="Row 10 not found"):
            await write_back(
                row_number=10,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

        # File should be unchanged
        with open(sample_csv) as f:
            assert f.read() == original_content

        # No temp files should remain
        dir_path = os.path.dirname(sample_csv)
        temp_files = [f for f in os.listdir(dir_path) if f.endswith(".csv.tmp")]
        assert len(temp_files) == 0


class TestWriteBackExcel:
    """Tests for Excel write-back functionality."""

    @pytest.mark.asyncio
    async def test_write_back_excel_adds_columns(self, sample_excel, mock_ctx):
        """Test that tracking columns are added when missing."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "excel",
            "path": sample_excel,
            "sheet": "Orders",
            "row_count": 3,
        }

        result = await write_back(
            row_number=1,
            tracking_number="1Z999AA10123456784",
            ctx=mock_ctx,
            shipped_at="2026-01-25T12:00:00Z",
        )

        assert result["success"] is True
        assert result["source_type"] == "excel"

        # Verify columns were added
        wb = load_workbook(sample_excel)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "tracking_number" in headers
        assert "shipped_at" in headers

        # Verify correct row was updated (row 1 = data row 1 = Excel row 2)
        tracking_col = headers.index("tracking_number") + 1
        shipped_col = headers.index("shipped_at") + 1
        assert ws.cell(row=2, column=tracking_col).value == "1Z999AA10123456784"
        assert ws.cell(row=2, column=shipped_col).value == "2026-01-25T12:00:00Z"
        wb.close()

    @pytest.mark.asyncio
    async def test_write_back_excel_updates_existing(
        self, sample_excel_with_tracking, mock_ctx
    ):
        """Test updating existing tracking columns."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "excel",
            "path": sample_excel_with_tracking,
            "sheet": "Orders",
            "row_count": 3,
        }

        result = await write_back(
            row_number=2,
            tracking_number="1Z999AA10123456785",
            ctx=mock_ctx,
            shipped_at="2026-01-25T14:00:00Z",
        )

        assert result["success"] is True

        # Verify row 2 was updated (Excel row 3)
        wb = load_workbook(sample_excel_with_tracking)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        tracking_col = headers.index("tracking_number") + 1
        shipped_col = headers.index("shipped_at") + 1

        assert ws.cell(row=3, column=tracking_col).value == "1Z999AA10123456785"
        assert ws.cell(row=3, column=shipped_col).value == "2026-01-25T14:00:00Z"
        # Row 1 and 3 should be unchanged
        assert ws.cell(row=2, column=tracking_col).value is None
        assert ws.cell(row=4, column=tracking_col).value is None
        wb.close()

    @pytest.mark.asyncio
    async def test_write_back_excel_correct_row(self, sample_excel, mock_ctx):
        """Test that row_number maps correctly to Excel row."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "excel",
            "path": sample_excel,
            "sheet": "Orders",
            "row_count": 3,
        }

        # Update row 3 (last data row)
        await write_back(
            row_number=3,
            tracking_number="1Z999AA10123456786",
            ctx=mock_ctx,
        )

        # Row 3 = Excel row 4 (header is row 1)
        wb = load_workbook(sample_excel)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        tracking_col = headers.index("tracking_number") + 1

        assert ws.cell(row=4, column=tracking_col).value == "1Z999AA10123456786"
        # Other rows unchanged
        assert ws.cell(row=2, column=tracking_col).value is None
        assert ws.cell(row=3, column=tracking_col).value is None
        wb.close()

    @pytest.mark.asyncio
    async def test_write_back_excel_sheet_selection(self, mock_ctx):
        """Test write-back works with named sheets."""
        # Create Excel with multiple sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["note"])
        ws1.append(["This is the summary"])

        ws2 = wb.create_sheet("Orders")
        ws2.append(["order_id", "customer"])
        ws2.append([1001, "John"])
        ws2.append([1002, "Jane"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            excel_path = f.name

        try:
            mock_ctx.request_context.lifespan_context["current_source"] = {
                "type": "excel",
                "path": excel_path,
                "sheet": "Orders",
                "row_count": 2,
            }

            await write_back(
                row_number=1,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

            # Verify the correct sheet was updated
            wb = load_workbook(excel_path)
            ws = wb["Orders"]
            headers = [cell.value for cell in ws[1]]
            assert "tracking_number" in headers

            tracking_col = headers.index("tracking_number") + 1
            assert ws.cell(row=2, column=tracking_col).value == "1Z999AA10123456784"

            # Summary sheet should be unchanged
            ws_summary = wb["Summary"]
            summary_headers = [cell.value for cell in ws_summary[1]]
            assert "tracking_number" not in summary_headers
            wb.close()
        finally:
            Path(excel_path).unlink(missing_ok=True)


class TestWriteBackDatabase:
    """Tests for database write-back functionality."""

    @pytest.mark.asyncio
    async def test_write_back_database_updates_row(self, mock_ctx, duckdb_conn):
        """Test that UPDATE executes correctly via DuckDB."""
        # Setup: Create table with identity tracking column
        duckdb_conn.execute(f"""
            CREATE TABLE orders (
                {SOURCE_ROW_NUM_COLUMN} INTEGER,
                order_id INTEGER,
                customer VARCHAR,
                tracking_number VARCHAR,
                shipped_at VARCHAR
            )
        """)
        duckdb_conn.execute("""
            INSERT INTO orders VALUES
            (1, 1001, 'John', NULL, NULL),
            (2, 1002, 'Jane', NULL, NULL)
        """)

        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "database",
            "query": "SELECT * FROM orders",
            "row_count": 2,
        }
        mock_ctx.request_context.lifespan_context["db"] = duckdb_conn

        result = await write_back(
            row_number=1,
            tracking_number="1Z999AA10123456784",
            ctx=mock_ctx,
            shipped_at="2026-01-25T12:00:00Z",
        )

        assert result["success"] is True
        assert result["source_type"] == "database"

        # Verify row was updated
        row = duckdb_conn.execute(
            f"SELECT tracking_number, shipped_at FROM orders WHERE {SOURCE_ROW_NUM_COLUMN} = 1"
        ).fetchone()
        assert row[0] == "1Z999AA10123456784"
        assert row[1] == "2026-01-25T12:00:00Z"

        # Other row unchanged
        row2 = duckdb_conn.execute(
            f"SELECT tracking_number FROM orders WHERE {SOURCE_ROW_NUM_COLUMN} = 2"
        ).fetchone()
        assert row2[0] is None

    @pytest.mark.asyncio
    async def test_write_back_database_parameterized(self, mock_ctx, duckdb_conn):
        """Test that queries use parameterization (no SQL injection)."""
        # Setup
        duckdb_conn.execute(f"""
            CREATE TABLE orders (
                {SOURCE_ROW_NUM_COLUMN} INTEGER,
                tracking_number VARCHAR,
                shipped_at VARCHAR
            )
        """)
        duckdb_conn.execute("INSERT INTO orders VALUES (1, NULL, NULL)")

        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "database",
            "query": "SELECT * FROM orders",
            "row_count": 1,
        }
        mock_ctx.request_context.lifespan_context["db"] = duckdb_conn

        # Attempt injection via tracking number
        malicious_tracking = "1Z'; DROP TABLE orders; --"

        result = await write_back(
            row_number=1,
            tracking_number=malicious_tracking,
            ctx=mock_ctx,
        )

        assert result["success"] is True

        # Table should still exist (no injection)
        count = duckdb_conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
        assert count == 1

        # Value should be stored literally
        row = duckdb_conn.execute(
            f"SELECT tracking_number FROM orders WHERE {SOURCE_ROW_NUM_COLUMN} = 1"
        ).fetchone()
        assert row[0] == malicious_tracking


class TestWriteBackErrors:
    """Tests for error handling in write_back."""

    @pytest.mark.asyncio
    async def test_write_back_no_source_loaded(self, mock_ctx):
        """Test error when no source is loaded."""
        mock_ctx.request_context.lifespan_context["current_source"] = None

        with pytest.raises(ValueError, match="No data source loaded"):
            await write_back(
                row_number=1,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

    @pytest.mark.asyncio
    async def test_write_back_unsupported_type(self, mock_ctx):
        """Test error for unknown source types."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "unknown_type",
            "path": "/some/path",
        }

        with pytest.raises(ValueError, match="Unsupported source type"):
            await write_back(
                row_number=1,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

    @pytest.mark.asyncio
    async def test_write_back_csv_row_not_found(self, sample_csv, mock_ctx):
        """Test error for non-existent row in CSV."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "csv",
            "path": sample_csv,
            "row_count": 3,
        }

        with pytest.raises(ValueError, match="Row 10 not found"):
            await write_back(
                row_number=10,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

    @pytest.mark.asyncio
    async def test_write_back_excel_row_not_found(self, sample_excel, mock_ctx):
        """Test error for non-existent row in Excel."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "excel",
            "path": sample_excel,
            "sheet": "Orders",
            "row_count": 3,
        }

        with pytest.raises(ValueError, match="Row 10 not found"):
            await write_back(
                row_number=10,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

    @pytest.mark.asyncio
    async def test_write_back_excel_sheet_not_found(self, sample_excel, mock_ctx):
        """Test error when sheet doesn't exist."""
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "excel",
            "path": sample_excel,
            "sheet": "NonExistent",
            "row_count": 3,
        }

        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            await write_back(
                row_number=1,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )

    @pytest.mark.asyncio
    async def test_write_back_database_complex_query(self, mock_ctx, duckdb_conn):
        """Test error for complex queries without extractable table."""
        # Use a subquery which cannot be written back to
        mock_ctx.request_context.lifespan_context["current_source"] = {
            "type": "database",
            "query": "SELECT * FROM (SELECT * FROM orders WHERE status = 'active') subq",
            "row_count": 1,
        }
        mock_ctx.request_context.lifespan_context["db"] = duckdb_conn

        with pytest.raises(ValueError, match="Cannot determine target table"):
            await write_back(
                row_number=1,
                tracking_number="1Z999AA10123456784",
                ctx=mock_ctx,
            )


class TestExtractTableName:
    """Tests for the _extract_table_name helper function."""

    def test_simple_select(self):
        """Test extraction from simple SELECT."""
        assert _extract_table_name("SELECT * FROM orders") == "orders"

    def test_select_with_where(self):
        """Test extraction with WHERE clause."""
        assert _extract_table_name("SELECT * FROM orders WHERE id = 1") == "orders"

    def test_select_with_schema(self):
        """Test extraction with schema prefix."""
        assert _extract_table_name("SELECT * FROM public.orders") == "public.orders"

    def test_select_with_columns(self):
        """Test extraction with specific columns."""
        assert _extract_table_name("SELECT id, name FROM customers") == "customers"

    def test_lowercase_from(self):
        """Test case-insensitive FROM."""
        assert _extract_table_name("select * from orders") == "orders"

    def test_no_from_clause(self):
        """Test query without FROM."""
        assert _extract_table_name("SELECT 1") is None

    def test_empty_query(self):
        """Test empty query string."""
        assert _extract_table_name("") is None

    def test_join_query(self):
        """Test that JOIN queries don't extract first table."""
        # JOIN queries should return the first table, but we document limitation
        result = _extract_table_name("SELECT * FROM orders JOIN customers ON orders.id = customers.order_id")
        # This will return "orders" - simple extraction works for first table
        assert result == "orders"
